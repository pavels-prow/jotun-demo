#!/usr/bin/env python3
import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def find_sheet_case_insensitive(wb, name):
    target = name.strip().lower()
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == target:
            return wb[sheet_name]
    return None


def get_used_range(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row <= 1 and max_col <= 1 and ws["A1"].value is None:
        return 0, 0
    return max_row, max_col


def clear_values(ws, max_row, max_col):
    if max_row <= 0 or max_col <= 0:
        return
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def copy_values_and_formats(src_ws, dst_ws, max_row, max_col):
    if max_row <= 0 or max_col <= 0:
        return
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_cell = src_ws.cell(row=r, column=c)
            dst_cell = dst_ws.cell(row=r, column=c)
            dst_cell.value = src_cell.value
            dst_cell.number_format = src_cell.number_format


def build_output_path(outdir, prefix):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if prefix:
        filename = f"{prefix}_{timestamp}.xlsx"
    else:
        filename = f"{timestamp}.xlsx"
    return Path(outdir) / filename


def parse_args(argv):
    parser = argparse.ArgumentParser(description="Merge data into an Excel template sheet.")
    parser.add_argument("--template", required=True, help="Path to template .xlsx")
    parser.add_argument("--data", required=True, help="Path to data .xlsx")
    parser.add_argument("--outdir", default=".", help="Output directory (default: current directory)")
    parser.add_argument("--sheet", default="data", help="Target sheet name in template (default: data)")
    parser.add_argument("--prefix", default="", help="Output filename prefix (default: empty)")
    parser.add_argument("--dry-run", action="store_true", help="Validate and print actions without writing output")
    return parser.parse_args(argv)


def main(argv):
    args = parse_args(argv)

    template_path = Path(args.template)
    data_path = Path(args.data)
    outdir = Path(args.outdir)

    try:
        template_wb = load_workbook(template_path)
    except Exception as exc:
        print(f"Error: failed to open template workbook: {exc}", file=sys.stderr)
        return 1

    try:
        data_wb = load_workbook(data_path, data_only=False)
    except Exception as exc:
        print(f"Error: failed to open data workbook: {exc}", file=sys.stderr)
        return 1

    target_ws = find_sheet_case_insensitive(template_wb, args.sheet)
    if target_ws is None:
        print(f"Error: template sheet not found: {args.sheet}", file=sys.stderr)
        return 2

    if len(data_wb.sheetnames) == 0:
        print("Error: data workbook has no sheets", file=sys.stderr)
        return 3

    source_ws = data_wb.worksheets[0]

    tmpl_rows, tmpl_cols = get_used_range(target_ws)
    clear_values(target_ws, tmpl_rows, tmpl_cols)

    src_rows, src_cols = get_used_range(source_ws)
    copy_values_and_formats(source_ws, target_ws, src_rows, src_cols)

    out_path = build_output_path(outdir, args.prefix)

    if args.dry_run:
        print("Dry run: no output written.")
    else:
        outdir.mkdir(parents=True, exist_ok=True)
        try:
            template_wb.save(out_path)
        except Exception as exc:
            print(f"Error: failed to save output workbook: {exc}", file=sys.stderr)
            return 1

    print("Summary:")
    print(f"  Template: {template_path}")
    print(f"  Data: {data_path}")
    print(f"  Target sheet: {target_ws.title}")
    print(f"  Source sheet: {source_ws.title}")
    print(f"  Copied: {src_rows} rows x {src_cols} cols")
    print(f"  Output: {out_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
