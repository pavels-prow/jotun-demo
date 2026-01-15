from typing import List, Dict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


HEADERS = ["Id", "Timestamp", "Category", "Customer", "Amount", "Status"]


def export_to_xlsx(path: str, rows: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(HEADERS)
    for row in rows:
        ws.append([row[h] for h in HEADERS])

    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(str(header))
        for row in rows:
            value = row[header]
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    wb.save(path)
